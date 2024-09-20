import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from io import BytesIO

def extract_otdr_data(text, wavelengths):
    """
    Extracts OTDR data based on the selected wavelengths and their order.
    """
    # Define the pattern to extract span length and span loss values
    span_length_pattern = r'Span length\s*\(ft\)\s*[:\-]\s*([\d,.]+)'
    span_loss_pattern = r'Span loss\s*\(dB\)\s*[:\-]\s*([\d,.]+)\s+([\d,.]+)?([\d,.]+)?'  # Allow for up to 3 values

    # Initialize the data dictionary with None values for all keys
    data = {'Span Length (ft)': None}

    # Add keys based on the selected wavelengths
    for wavelength in wavelengths:
        data[f'Span Loss {wavelength}nm (dB)'] = None

    # Extract span length
    span_length_match = re.search(span_length_pattern, text, re.IGNORECASE)
    if span_length_match:
        data['Span Length (ft)'] = float(span_length_match.group(1).replace(',', ''))

    # Extract span loss for available wavelengths
    span_loss_match = re.search(span_loss_pattern, text, re.IGNORECASE)
    if span_loss_match:
        for i, wavelength in enumerate(wavelengths):
            if span_loss_match.group(i + 1):  # Check if the span loss value is available
                data[f'Span Loss {wavelength}nm (dB)'] = float(span_loss_match.group(i + 1).replace(',', ''))

    return data

def insert_data_into_template(template_path, data_df):
    """
    Inserts data from the DataFrame into the specified Excel template.
    Returns the updated Excel file in memory.
    """
    wb = load_workbook(template_path)
    
    sheet_name = 'OTDR_Data'
    start_cell = 'A2'
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the Excel template.")
    
    sheet = wb[sheet_name]
    
    # Convert start_cell to row and column indices
    match = re.match(r"([A-Za-z]+)(\d+)", start_cell)
    start_col_letter, start_row = match.groups()
    start_col = column_index_from_string(start_col_letter)
    start_row = int(start_row)
    
    # Insert data into the Excel template
    for idx, row in data_df.iterrows():
        for col_idx, value in enumerate(row, start=0):
            cell = sheet.cell(row=start_row + idx, column=start_col + col_idx)
            cell.value = value
    
    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
