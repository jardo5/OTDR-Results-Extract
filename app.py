import streamlit as st
import pdfplumber
import pandas as pd
import tempfile
import os
from openpyxl import load_workbook
from io import BytesIO
import re
import openpyxl

# --------------------------
# Helper Functions
# --------------------------

def extract_otdr_data(text, wavelengths):
    """
    Extracts OTDR data based on the selected wavelengths and their order.
    """
    # Define the pattern to extract span length and span loss values
    span_length_pattern = r'Span length\s*\(ft\)\s*[:\-]\s*([\d,.]+)'
    span_loss_pattern = r'Span loss\s*\(dB\)\s*[:\-]\s*([\d,.]+)\s+([\d,.]+)?([\d,.]+)?'  # Allow for up to 3 values

    # Initialize the data dictionary with None values for all keys
    data = {'Span Length (ft)': None}

    # Add keys based on the selected wavelengths
    for wavelength in wavelengths:
        data[f'Span Loss {wavelength}nm (dB)'] = None

    # Extract span length
    span_length_match = re.search(span_length_pattern, text, re.IGNORECASE)
    if span_length_match:
        data['Span Length (ft)'] = float(span_length_match.group(1).replace(',', ''))

    # Extract span loss for available wavelengths
    span_loss_match = re.search(span_loss_pattern, text, re.IGNORECASE)
    if span_loss_match:
        for i, wavelength in enumerate(wavelengths):
            if span_loss_match.group(i + 1):  # Check if the span loss value is available
                data[f'Span Loss {wavelength}nm (dB)'] = float(span_loss_match.group(i + 1).replace(',', ''))

    return data


def insert_data_into_template(template_path, data_df):
    """
    Inserts data from the DataFrame into the specified Excel template.
    Returns the updated Excel file in memory.
    """
    wb = load_workbook(template_path)
    
    sheet_name = 'OTDR_Data'
    start_cell = 'A2'
    
    if sheet_name not in wb.sheetnames:
        st.error(f"Sheet '{sheet_name}' not found in the Excel template.")
        return None
    
    sheet = wb[sheet_name]
    
    # Convert start_cell to row and column indices
    match = re.match(r"([A-Za-z]+)(\d+)", start_cell)
    start_col_letter, start_row = match.groups()
    start_col = openpyxl.utils.column_index_from_string(start_col_letter)
    start_row = int(start_row)
    
    # Insert data into the Excel template
    for idx, row in data_df.iterrows():
        for col_idx, value in enumerate(row, start=0):
            cell = sheet.cell(row=start_row + idx, column=start_col + col_idx)
            cell.value = value
    
    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --------------------------
# Streamlit App Layout
# --------------------------

image_path = "assets/schupps-logo.png"

st.set_page_config(page_title="SLC", page_icon=image_path, layout="wide")

left_co, cent_co, last_co = st.columns(3)
with cent_co:
    st.image(image_path, caption='FOR USE BY AUTHORIZED SLC EMPLOYEES', use_column_width=True)

st.title("OTDR Results Extractor to Excel")

st.markdown("""
Extract .SOR data from OTDR PDF reports and create easy copy and paste Excel files.

**How To:**
- Put in multiple PDF files :red[**IN ORDER**].
- :red[**SELECT THE CORRECT WAVELENGTHS**].
- Give it time to process the data.
- After processing, you can download the updated Excel file.
- Do as you will with the data, double check.

**Notes:**
- **Issues?** Don't call me Mike/Dom, call George.
- My new rate is $700/hr for fixes.
- :red[**1310 HASN'T BEEN FULLY TESTED YET**]
""")

# Allow user to select which wavelengths are present
wavelength_options = {
    "1310nm & 1550nm": [1310, 1550],
    "1550nm & 1625nm": [1550, 1625],
    "1310nm & 1625nm": [1310, 1625],
    "1310nm, 1550nm & 1625nm": [1310, 1550, 1625]
}

selected_wavelengths = st.selectbox(
    "Select Wavelengths Present in the PDF Data:",
    list(wavelength_options.keys())
)

wavelengths = wavelength_options[selected_wavelengths]

# File Uploaders
uploaded_pdfs = st.file_uploader("", type=["pdf"], accept_multiple_files=True)

# Maximum file size (e.g., 10MB)
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

# Process Button
if st.button("🔍 Extract and Populate Data"):
    if not uploaded_pdfs:
        st.error("Please upload at least one PDF file.")
    else:
        # Check file sizes
        oversized_files = [pdf.name for pdf in uploaded_pdfs if pdf.size > MAX_FILE_SIZE]
        if oversized_files:
            st.error(f"The following PDF files exceed the maximum allowed size of 10MB: {', '.join(oversized_files)}")
        else:
            try:
                # Select the appropriate template based on the chosen wavelengths
                template_paths = {
                    "1310nm & 1550nm": 'templates/OTDR_Template_1310_1550.xlsx',
                    "1550nm & 1625nm": 'templates/OTDR_Template_1550_1625.xlsx',
                    "1310nm & 1625nm": 'templates/OTDR_Template_1310_1625.xlsx',
                    "1310nm, 1550nm & 1625nm": 'templates/OTDR_Template_All.xlsx'
                }
                template_path = template_paths[selected_wavelengths]
            
                # Initialize list to collect all data
                all_data = []
                fiber_counter = 1  # To keep the fiber number continuous across multiple PDFs
            
                # Initialize progress bar
                total_pdfs = len(uploaded_pdfs)
                progress_bar = st.progress(0)
                current_pdf = 0
            
                for pdf_file in uploaded_pdfs:
                    current_pdf += 1
                    st.info(f"Processing {pdf_file.name} ({current_pdf}/{total_pdfs})...")
                    
                    # Save PDF to a temporary file
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
                        tmp_pdf.write(pdf_file.read())
                        pdf_path = tmp_pdf.name
                    
                    with pdfplumber.open(pdf_path) as pdf:
                        total_pages = len(pdf.pages)
                        for page_num, page in enumerate(pdf.pages, start=1):
                            text = page.extract_text()
                            if not text:
                                st.warning(f"No text found on page {page_num} of {pdf_file.name}. Skipping.")
                                continue
                            data = extract_otdr_data(text, wavelengths)
                            data['PDF File Name'] = pdf_file.name
                            data['Fiber #/Page #'] = fiber_counter  # Keep the fiber number continuous
                            fiber_counter += 1  # Increment the fiber number for the next page
                            all_data.append(data)
                    
                    # Remove temporary PDF file
                    os.remove(pdf_path)
                    
                    # Update progress bar
                    progress = current_pdf / total_pdfs
                    progress_bar.progress(progress)
            
                if not all_data:
                    st.error("No data extracted from the uploaded PDFs.")
                else:
                    # Create DataFrame
                    df = pd.DataFrame(all_data)
                    
                    # Reorder columns and handle missing Span Loss data
                    desired_columns = ['PDF File Name', 'Fiber #/Page #', 'Span Length (ft)'] + \
                                    [f'Span Loss {wl}nm (dB)' for wl in wavelengths]
                    
                    for col in desired_columns:
                        if col not in df.columns:
                            df[col] = None  # Assign None if column not present
                    df = df[desired_columns]
                    
                    st.success("Data extraction complete!")
                    st.subheader("Extracted Data Preview")
                    st.dataframe(df.head(20))
                    
                    # Insert data into Excel Template
                    st.info("Inserting data into the Excel template...")
                    with st.spinner('Populating Excel template...'):
                        updated_excel = insert_data_into_template(template_path, df)
                    
                    if updated_excel:
                        st.success("Data successfully inserted into the template!")
                        
                        # Provide download link
                        st.download_button(
                            label="Download Updated Excel",
                            data=updated_excel,
                            file_name="Updated_OTDR_Results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            
            except Exception as e:
                st.error(f"An error occurred: {e}")
